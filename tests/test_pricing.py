"""Рушій ціноутворення: джерела цін, регіональні коефіцієнти, підсумки."""
import pytest

from backend.services import catalog, pricing


def test_region_factor_prefers_city_over_region():
    assert catalog.region_factor("Київ", "Полтавська область")[0] == 1.35
    assert catalog.region_factor("", "Львівська область")[0] == 1.14
    assert catalog.region_factor("Лохвиця", "Полтавська область")[0] == 0.95
    assert catalog.region_factor("Невідоме місто", "")[0] == 1.0


def test_catalog_price_is_scaled_by_region(clean_db):
    name = "Демонтаж розеток"
    base = pricing.resolve_price(name, "шт", "", "")
    kyiv = pricing.resolve_price(name, "шт", "Київ", "")
    assert base.source == "catalog"
    assert kyiv.labor == pytest.approx(base.labor * 1.35, rel=1e-6)
    assert kyiv.region_label == "місто Київ"


def test_history_takes_priority_over_catalog(clean_db):
    name = "Демонтаж розеток"
    pricing.remember_price(name, "шт", "Полтава", "Полтавська область",
                           labor=222.0, material=11.0, machines=0.0)
    res = pricing.resolve_price(name, "шт", "Полтава", "Полтавська область")
    assert res.source == "history"
    assert res.labor == 222.0
    assert res.details["scope"] == "місто Полтава"


def test_history_is_scoped_to_city(clean_db):
    name = "Демонтаж розеток"
    pricing.remember_price(name, "шт", "Одеса", "Одеська область", 500.0, 0.0, 0.0)
    # Для іншого міста своєї історії немає — береться загальна.
    res = pricing.resolve_price(name, "шт", "Харків", "Харківська область")
    assert res.source == "history"
    assert res.details["scope"] == "уся історія"


def test_history_uses_median_of_recent_prices(clean_db):
    name = "Демонтаж світильників"
    for value in (100.0, 200.0, 900.0):
        pricing.remember_price(name, "шт", "Полтава", "Полтавська область", value, 0.0, 0.0)
    res = pricing.resolve_price(name, "шт", "Полтава", "Полтавська область")
    assert res.labor == 200.0   # медіана стійка до викиду 900


def test_catalog_only_strategy_ignores_history(clean_db):
    name = "Демонтаж розеток"
    pricing.remember_price(name, "шт", "Полтава", "Полтавська область", 999.0, 0.0, 0.0)
    res = pricing.resolve_price(name, "шт", "Полтава", "", strategy="catalog_only")
    assert res.source == "catalog"


def test_unknown_work_has_no_price(clean_db):
    res = pricing.resolve_price("Політ на Марс приватним шатлом", "шт", "Полтава", "")
    assert res.source == "none"
    assert res.unit_total == 0


def test_totals_chain():
    tree = [{"code": "01", "title": "К", "divisions": [{"code": "1", "title": "Р", "positions": [
        {"name": "робота", "quantity": 10, "labor_price": 100,
         "material_price": 50, "machines_price": 10},
    ]}]}]
    result = pricing.calculate(tree, {"overhead_pct": 15, "profit_pct": 8, "admin_pct": 2,
                                      "risk_pct": 0, "vat_pct": 20, "materials_included": True})
    assert result["direct"]["total"] == 1600.0     # (100+50+10) * 10
    assert result["direct"]["labor"] == 1000.0
    # ЗВВ рахуються від робіт і машин, без матеріалів: (1000 + 100) * 15%
    assert result["overhead"] == 165.0
    assert result["profit"] == pytest.approx((1100 + 165) * 0.08)
    assert result["vat"] == pytest.approx(result["before_vat"] * 0.2)
    assert result["grand_total"] == pytest.approx(result["before_vat"] * 1.2)


def test_materials_can_be_excluded():
    tree = [{"code": "", "title": "", "divisions": [{"code": "", "title": "", "positions": [
        {"name": "робота", "quantity": 2, "labor_price": 100,
         "material_price": 900, "machines_price": 0},
    ]}]}]
    result = pricing.calculate(tree, {"materials_included": False})
    assert result["direct"]["material"] == 0.0
    assert result["direct"]["total"] == 200.0


def test_export_keeps_full_quantity_precision(clean_db):
    """Обсяг 9,690009 т не має перетворюватись на 9,69 у вивантаженні."""
    from backend.services import exporter

    calc = {
        "object": {"name": "Тест", "address": "", "city": "Полтава", "region": "",
                   "customer": "", "doc_code": ""},
        "estimates": [{"code": "01", "title": "К", "totals": {"labor": 0, "material": 0,
                                                              "machines": 0, "total": 0},
                       "divisions": [{"code": "1", "title": "Р",
                                      "totals": {"labor": 0, "material": 0,
                                                 "machines": 0, "total": 0},
                                      "positions": [{
                                          "number": 1, "name": "Навантаження сміття вручну",
                                          "unit": "т", "quantity": 9.690009,
                                          "labor_price": 380.0, "material_price": 0.0,
                                          "machines_price": 0.0, "price_source": "catalog",
                                          "totals": {"unit_price": 380.0, "labor_total": 3682.2,
                                                     "material_total": 0.0, "machines_total": 0.0,
                                                     "total": 3682.2}}]}]}],
        "lines": [{"key": "total", "label": "Усього", "value": 3682.2}],
        "settings": {},
    }
    assert exporter._position_cells(calc["estimates"][0]["divisions"][0]["positions"][0])[3] \
        == 9.690009
    assert "9,690009" in exporter.to_html(calc) or "9.690009" in exporter.to_html(calc)

    import io

    from openpyxl import load_workbook
    ws = load_workbook(io.BytesIO(exporter.to_xlsx(calc))).active
    quantities = [ws.cell(row=r, column=4).value for r in range(1, ws.max_row + 1)]
    assert 9.690009 in quantities
