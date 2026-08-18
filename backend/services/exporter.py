"""Вивантаження заповненого кошторису: XLSX, PDF та HTML для друку."""
from __future__ import annotations

import html
import io
from datetime import date

from .pricing import SOURCE_LABELS

TITLE = "Локальний кошторис"
# (повна назва колонки, ширина в XLSX, коротка назва для вузьких колонок PDF)
COLUMNS = [
    ("№\nп/п", 6, "№"),
    ("Найменування робіт та витрат", 62, "Найменування робіт та витрат"),
    ("Одиниця\nвиміру", 10, "Од.\nвим."),
    ("Кількість", 11, "К-сть"),
    ("Ціна за одиницю, грн\nроботи", 13, "Ціна/од.\nроботи"),
    ("Ціна за одиницю, грн\nматеріали", 13, "Ціна/од.\nматер."),
    ("Ціна за одиницю, грн\nмашини", 12, "Ціна/од.\nмашини"),
    ("Ціна за одиницю, грн\nвсього", 13, "Ціна/од.\nвсього"),
    ("Вартість, грн\nроботи", 14, "Вартість\nроботи"),
    ("Вартість, грн\nматеріали", 14, "Вартість\nматеріали"),
    ("Вартість, грн\nмашини", 13, "Вартість\nмашини"),
    ("Вартість, грн\nвсього", 15, "Вартість\nвсього"),
    ("Джерело\nціни", 16, "Джерело\nціни"),
]


def _money(value) -> float:
    try:
        return round(float(value or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def _qty(value) -> float:
    """Обсяг не можна округляти до копійок: у відомостях трапляється 9,690009 т."""
    try:
        return round(float(value or 0), 6)
    except (TypeError, ValueError):
        return 0.0


def _fmt_qty(value, separator: str = " ") -> str:
    """Обсяг без зайвих нулів: 9,690009 → «9,690009», 4 → «4»."""
    text = f"{_qty(value):,.6f}".rstrip("0").rstrip(".")
    return (text or "0").replace(",", separator)


def _source_label(position: dict) -> str:
    return SOURCE_LABELS.get(position.get("price_source", ""), position.get("price_source", "") or "—")


def _rows(calc: dict):
    """Плаский потік рядків для будь-якого формату: (тип, дані)."""
    for estimate in calc["estimates"]:
        title = " ".join(x for x in [estimate.get("code", ""), estimate.get("title", "")] if x)
        yield "estimate", {"title": f"{TITLE} {title}".strip(), "totals": estimate.get("totals", {})}
        for division in estimate["divisions"]:
            code = division.get("code", "")
            head = f"Розділ {code}. {division.get('title', '')}" if code else division.get("title", "")
            if head.strip():
                yield "division", {"title": head.strip(), "totals": division.get("totals", {})}
            for position in division["positions"]:
                yield "position", position
            if head.strip():
                yield "division_total", {"title": f"Разом по розділу «{division.get('title', '')}»",
                                        "totals": division.get("totals", {})}
        yield "estimate_total", {"title": f"Разом по кошторису {title}".strip(),
                                 "totals": estimate.get("totals", {})}


def _position_cells(position: dict) -> list:
    totals = position.get("totals", {})
    return [
        position.get("number", ""),
        position.get("name", ""),
        position.get("unit", ""),
        _qty(position.get("quantity")),
        _money(position.get("labor_price")),
        _money(position.get("material_price")),
        _money(position.get("machines_price")),
        _money(totals.get("unit_price")),
        _money(totals.get("labor_total")),
        _money(totals.get("material_total")),
        _money(totals.get("machines_total")),
        _money(totals.get("total")),
        _source_label(position),
    ]


def _header_lines(calc: dict) -> list[str]:
    obj = calc["object"]
    lines = [obj.get("name", "")]
    if obj.get("address"):
        lines.append(f"Адреса: {obj['address']}")
    place = ", ".join(x for x in [obj.get("city", ""), obj.get("region", "")] if x)
    if place:
        lines.append(f"Місце будівництва: {place}")
    if obj.get("customer"):
        lines.append(f"Замовник: {obj['customer']}")
    if obj.get("doc_code"):
        lines.append(f"Шифр: {obj['doc_code']}")
    lines.append(f"Складено станом на {date.today().strftime('%d.%m.%Y')}")
    return [x for x in lines if x]


# ------------------------------------------------------------------------- XLSX

def to_xlsx(calc: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Кошторис"

    thin = Side(style="thin", color="B0B7C3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="1F3864")
    est_fill = PatternFill("solid", fgColor="D9E2F3")
    div_fill = PatternFill("solid", fgColor="EDF2FA")
    total_fill = PatternFill("solid", fgColor="FFF2CC")
    money_fmt = "#,##0.00"
    qty_fmt = "#,##0.######"   # обсяги бувають з 4–6 знаками (напр. 9,690009 т)

    row = 1
    for line in _header_lines(calc):
        ws.cell(row=row, column=1, value=line).font = Font(bold=(row == 1), size=12 if row == 1 else 10)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
        row += 1
    row += 1

    header_row = row
    for idx, (title, width, _short) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=idx, value=title.replace("\n", " · "))
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = head_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[header_row].height = 34
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    row += 1

    for kind, data in _rows(calc):
        if kind == "position":
            for idx, value in enumerate(_position_cells(data), start=1):
                cell = ws.cell(row=row, column=idx, value=value)
                cell.border = border
                cell.font = Font(size=9)
                if idx == 2:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                elif idx >= 4:
                    cell.number_format = qty_fmt if idx == 4 else money_fmt
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="top")
            row += 1
            continue

        totals = data.get("totals", {})
        cell = ws.cell(row=row, column=1, value=data["title"])
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        fill = {"estimate": est_fill, "division": div_fill}.get(kind, total_fill)
        cell.font = Font(bold=True, size=10 if kind == "estimate" else 9)
        for idx in range(1, len(COLUMNS) + 1):
            c = ws.cell(row=row, column=idx)
            c.fill = fill
            c.border = border
        if kind in ("division_total", "estimate_total"):
            for offset, key in enumerate(("labor", "material", "machines", "total")):
                c = ws.cell(row=row, column=9 + offset, value=_money(totals.get(key)))
                c.number_format = money_fmt
                c.font = Font(bold=True, size=9)
                c.alignment = Alignment(horizontal="right")
                c.fill = fill
                c.border = border
        row += 1

    row += 1
    for line in calc["lines"]:
        label = ws.cell(row=row, column=1, value=("    " if line.get("sub") else "") + line["label"])
        label.font = Font(bold=not line.get("sub"), size=10)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
        value = ws.cell(row=row, column=12, value=_money(line["value"]))
        value.number_format = money_fmt
        value.font = Font(bold=not line.get("sub"), size=10)
        value.alignment = Alignment(horizontal="right")
        if line["key"] == "total":
            for idx in range(1, len(COLUMNS) + 1):
                ws.cell(row=row, column=idx).fill = total_fill
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Склав ____________________  [посада, підпис, прізвище]").font = Font(size=9)
    ws.cell(row=row + 2, column=1, value="Перевірив ________________  [посада, підпис, прізвище]").font = Font(size=9)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# -------------------------------------------------------------------------- PDF

_FONT_CANDIDATES = [
    ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
    ("/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
     "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf"),
    ("/usr/share/fonts/truetype/freefont/FreeSans.ttf", "/usr/share/fonts/truetype/freefont/FreeSansBold.ttf"),
]


def _register_font() -> tuple[str, str]:
    """Реєструє шрифт із кирилицею; повертає назви звичайного та жирного накреслення."""
    import os

    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    for regular, bold in _FONT_CANDIDATES:
        if os.path.exists(regular) and os.path.exists(bold):
            if "BudsmetSans" not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont("BudsmetSans", regular))
                pdfmetrics.registerFont(TTFont("BudsmetSans-Bold", bold))
            return "BudsmetSans", "BudsmetSans-Bold"
    return "Helvetica", "Helvetica-Bold"


def to_pdf(calc: dict) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (KeepTogether, Paragraph, SimpleDocTemplate,
                                    Spacer, Table, TableStyle)

    font, font_bold = _register_font()
    title_style = ParagraphStyle("t", fontName=font_bold, fontSize=11, leading=14, alignment=TA_CENTER)
    meta_style = ParagraphStyle("m", fontName=font, fontSize=8.5, leading=11, alignment=TA_CENTER)
    cell_style = ParagraphStyle("c", fontName=font, fontSize=7, leading=8.5)
    head_style = ParagraphStyle("h", fontName=font_bold, fontSize=7, leading=8.5,
                                alignment=TA_CENTER, textColor=colors.white)
    band_style = ParagraphStyle("b", fontName=font_bold, fontSize=8, leading=10)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=10 * mm, rightMargin=10 * mm,
                            topMargin=10 * mm, bottomMargin=10 * mm,
                            title=calc["object"].get("name", TITLE))
    story = [Paragraph(html.escape(TITLE.upper()), title_style), Spacer(1, 3 * mm)]
    for line in _header_lines(calc):
        story.append(Paragraph(html.escape(line), meta_style))
    story.append(Spacer(1, 4 * mm))

    widths = [9, 104, 16, 21, 23, 23, 21, 23, 26, 26, 23, 28, 24]
    scale = (doc.width) / sum(widths)
    widths = [w * scale for w in widths]

    head = [Paragraph(html.escape(short).replace("\n", "<br/>"), head_style)
            for _t, _w, short in COLUMNS]
    data = [head]
    styles = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3864")),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#B0B7C3")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (3, 1), (-2, -1), "RIGHT"),
        ("ALIGN", (0, 1), (0, -1), "CENTER"),
        ("ALIGN", (2, 1), (2, -1), "CENTER"),
        ("FONTNAME", (0, 1), (-1, -1), font),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]

    def fmt(value) -> str:
        return f"{_money(value):,.2f}".replace(",", " ")

    idx = 1
    for kind, item in _rows(calc):
        if kind == "position":
            cells = _position_cells(item)
            data.append([str(cells[0]), Paragraph(html.escape(str(cells[1])), cell_style),
                         str(cells[2]), _fmt_qty(cells[3])] +
                        [fmt(v) for v in cells[4:12]] +
                        [Paragraph(html.escape(str(cells[12])), cell_style)])
        else:
            totals = item.get("totals", {})
            band = [Paragraph(html.escape(item["title"]), band_style)] + [""] * (len(COLUMNS) - 1)
            if kind in ("division_total", "estimate_total"):
                for offset, key in enumerate(("labor", "material", "machines", "total")):
                    band[8 + offset] = fmt(totals.get(key))
            data.append(band)
            color = {"estimate": "#D9E2F3", "division": "#EDF2FA"}.get(kind, "#FFF2CC")
            styles += [("BACKGROUND", (0, idx), (-1, idx), colors.HexColor(color)),
                       ("SPAN", (0, idx), (7, idx)),
                       ("FONTNAME", (0, idx), (-1, idx), font_bold)]
        idx += 1

    table = Table(data, colWidths=widths, repeatRows=1)
    table.setStyle(TableStyle(styles))
    story.append(table)
    story.append(Spacer(1, 5 * mm))

    summary = [[Paragraph(("    " if line.get("sub") else "") + html.escape(line["label"]),
                          cell_style if line.get("sub") else band_style), fmt(line["value"])]
               for line in calc["lines"]]
    summary_table = Table(summary, colWidths=[doc.width * 0.75, doc.width * 0.25])
    summary_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#B0B7C3")),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTNAME", (1, 0), (1, -1), font_bold),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, len(summary) - 1), (-1, len(summary) - 1), colors.HexColor("#FFF2CC")),
    ]))
    story.append(KeepTogether(summary_table))
    story.append(Spacer(1, 8 * mm))
    story.append(Paragraph("Склав ____________________  [посада, підпис, прізвище]", cell_style))
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph("Перевірив ________________  [посада, підпис, прізвище]", cell_style))

    doc.build(story)
    return buffer.getvalue()


# -------------------------------------------------------------------------- HTML

def to_html(calc: dict) -> str:
    """Версія для друку / збереження в PDF засобами браузера."""
    def esc(value) -> str:
        return html.escape(str(value if value is not None else ""))

    def fmt(value) -> str:
        return f"{_money(value):,.2f}".replace(",", "&nbsp;")

    head_cells = "".join(f"<th>{esc(t).replace(chr(10), '<br>')}</th>" for t, _w, _s in COLUMNS)
    body = []
    for kind, item in _rows(calc):
        if kind == "position":
            cells = _position_cells(item)
            tds = [f"<td class=num>{esc(cells[0])}</td>", f"<td>{esc(cells[1])}</td>",
                   f"<td class=num>{esc(cells[2])}</td>"]
            tds.append(f"<td class=money>{_fmt_qty(cells[3], '&nbsp;')}</td>")
            tds += [f"<td class=money>{fmt(v)}</td>" for v in cells[4:12]]
            tds.append(f"<td class=src>{esc(cells[12])}</td>")
            body.append("<tr>" + "".join(tds) + "</tr>")
        else:
            totals = item.get("totals", {})
            money = ""
            if kind in ("division_total", "estimate_total"):
                money = "".join(f"<td class=money>{fmt(totals.get(k))}</td>"
                                for k in ("labor", "material", "machines", "total"))
            else:
                money = "<td colspan=4></td>"
            body.append(f"<tr class='band {kind}'><td colspan=8>{esc(item['title'])}</td>"
                        f"{money}<td></td></tr>")

    summary = "".join(
        f"<tr class='{'sub' if line.get('sub') else ''} {'grand' if line['key'] == 'total' else ''}'>"
        f"<td>{esc(line['label'])}</td><td class=money>{fmt(line['value'])}</td></tr>"
        for line in calc["lines"])

    header = "".join(f"<div>{esc(line)}</div>" for line in _header_lines(calc))
    return f"""<!doctype html>
<html lang="uk"><head><meta charset="utf-8">
<title>{esc(calc['object'].get('name', TITLE))}</title>
<style>
 @page {{ size: A4 landscape; margin: 10mm; }}
 body {{ font: 11px/1.35 "Segoe UI", Arial, sans-serif; color: #111; }}
 h1 {{ font-size: 15px; text-align: center; margin: 0 0 6px; }}
 .meta {{ text-align: center; font-size: 11px; margin-bottom: 12px; color: #333; }}
 table {{ border-collapse: collapse; width: 100%; }}
 th, td {{ border: 1px solid #b0b7c3; padding: 3px 4px; vertical-align: top; }}
 th {{ background: #1f3864; color: #fff; font-size: 10px; text-align: center; }}
 td.money {{ text-align: right; white-space: nowrap; }}
 td.num, td.src {{ text-align: center; font-size: 10px; }}
 tr.band td {{ font-weight: 700; background: #edf2fa; }}
 tr.estimate td {{ background: #d9e2f3; }}
 tr.division_total td, tr.estimate_total td {{ background: #fff2cc; }}
 .summary {{ margin-top: 14px; width: 60%; margin-left: auto; }}
 .summary td {{ padding: 4px 6px; }}
 .summary tr.sub td {{ color: #555; font-size: 10px; padding-left: 18px; }}
 .summary tr.grand td {{ background: #fff2cc; font-weight: 700; font-size: 13px; }}
 .sign {{ margin-top: 24px; font-size: 11px; }}
</style></head><body>
<h1>{esc(TITLE.upper())}</h1>
<div class="meta">{header}</div>
<table><thead><tr>{head_cells}</tr></thead><tbody>{''.join(body)}</tbody></table>
<table class="summary">{summary}</table>
<div class="sign">Склав ____________________ [посада, підпис, прізвище]</div>
<div class="sign">Перевірив ________________ [посада, підпис, прізвище]</div>
</body></html>"""
