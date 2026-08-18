"""Імпорт порожньої «Відомості обсягів робіт» (АВК-5) з PDF / XLSX / CSV.

Результат — `ImportedDocument`: шапка об'єкта плюс плаский перелік секцій
(локальні кошториси та розділи) з позиціями «найменування / од.вим. / кількість».
"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import asdict, dataclass, field

from .normalize import clean_text, normalize_unit, parse_number, soft_clean

# Одиниці виміру, що реально трапляються у відомостях обсягів робіт.
KNOWN_UNITS = {
    "шт", "м", "м2", "м3", "т", "кг", "к-т", "компл", "комплект", "блок",
    "грати", "агрегат", "квт", "кВт", "кінців", "лінія", "вимір.", "вимір",
    "випроб.", "випроб", "точ.", "точ", "струм-ч", "фазув-ня", "вимірюв.",
    "вимірюв", "м.п", "мп", "%", "маш-год", "люд-год", "місце", "система",
    "пара", "секція", "прилад", "апарат", "пристрій", "вузол", "проба",
    "фазув-ня.", "од", "штук", "пог.м", "кв.м", "куб.м", "канал", "стояк",
}
_KNOWN_UNITS_NORM = {normalize_unit(u) for u in KNOWN_UNITS}

_ESTIMATE_RE = re.compile(r"^Локальн\w*\s+кошторис\b(?P<rest>.*)$", re.I)
_DIVISION_RE = re.compile(r"^Розд[іi]л\b\s*(?P<num>\d+)?\s*[.:)]?\s*(?P<title>.*)$", re.I)
_POSITION_START_RE = re.compile(r"^(?P<num>\d{1,4})\s+(?P<rest>\D.*)$")
_NOISE_RE = re.compile(
    r"Програмний комплекс|^\s*\d+\s*$|^\s*1\s+2\s+3\s+4\s+5\s*$|"
    r"^Додаток\s|^до ДБН|^затверджений|^Мінрегіонбуд|^від \d{2}\.\d{2}\.\d{4}|"
    r"^\s*№\s*$|^п/п\s|^вимір[уy]\s|^\[посада"
)
_DOC_CODE_RE = re.compile(r"\b(\d{2,5}[_-][А-ЯЇІЄҐA-Z]{1,5}[_-][А-ЯЇІЄҐA-Z]{2,6})\b")
_HEADER_TITLE_RE = re.compile(r"^Відомість\s+обсяг", re.I)
_TABLE_HEAD_RE = re.compile(r"Найменування\s+робіт|^\s*№\s*$|^п/п\b")
_OBJECT_HINT_RE = re.compile(r"(ремонт|будівництв|реконструкц|реставрац|облаштув|благоустр)", re.I)

# Хвіст рядка позиції: «…назва   м2 151,1» (колонки розділені 2+ пробілами).
_TAIL_STRICT_RE = re.compile(r"\s{2,}(?P<unit>\S+(?:\s\S+)?)\s+(?P<qty>-?\d[\d ]*(?:[.,]\d+)?)\s*$")
_TAIL_LOOSE_RE = re.compile(r"\s(?P<unit>[^\s\d][^\s]{0,11})\s+(?P<qty>-?\d[\d ]*(?:[.,]\d+)?)\s*$")


@dataclass
class ImportedPosition:
    number: int
    name: str
    unit: str
    quantity: float
    note: str = ""


@dataclass
class ImportedSection:
    kind: str  # "estimate" — локальний кошторис, "division" — розділ
    code: str = ""
    title: str = ""
    positions: list = field(default_factory=list)


@dataclass
class ImportedDocument:
    object_name: str = ""
    address: str = ""
    city: str = ""
    region: str = ""
    doc_code: str = ""
    sections: list = field(default_factory=list)
    warnings: list = field(default_factory=list)

    @property
    def positions(self):
        return [p for s in self.sections for p in s.positions]

    def to_dict(self):
        data = asdict(self)
        data["total_positions"] = len(self.positions)
        return data


# --------------------------------------------------------------------------- PDF

def _pdf_lines(data: bytes) -> tuple[list[str], str]:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    lines: list[str] = []
    doc_code = ""
    for page in reader.pages:
        for raw in (page.extract_text() or "").split("\n"):
            line = soft_clean(raw)
            if not line.strip():
                continue
            if "Програмний комплекс" in line:
                hit = _DOC_CODE_RE.search(line)
                if hit and not doc_code:
                    doc_code = hit.group(1)
                continue
            if _NOISE_RE.search(line.strip()):
                continue
            lines.append(line)
    return lines, doc_code


def parse_pdf(data: bytes) -> ImportedDocument:
    lines, doc_code = _pdf_lines(data)
    doc = _parse_lines(lines)
    doc.doc_code = doc_code or doc.doc_code
    return doc


# -------------------------------------------------------------------------- XLSX

def _xlsx_rows(data: bytes) -> list[list[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    rows: list[list[str]] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            cells = ["" if c is None else clean_text(c) for c in row]
            if any(cells):
                rows.append(cells)
    wb.close()
    return rows


def _parse_table_rows(rows: list[list[str]]) -> ImportedDocument:
    """Табличні джерела (XLSX/CSV) — колонки вже розділені, шукаємо шапку таблиці."""
    doc = ImportedDocument()
    current: ImportedSection | None = None
    header: dict[str, int] | None = None
    counter = 0

    for cells in rows:
        joined = clean_text(" ".join(c for c in cells if c))
        if not joined:
            continue

        if header is None:
            low = [c.lower() for c in cells]
            if any("найменув" in c for c in low) and any(
                ("кільк" in c) or ("обсяг" in c) for c in low
            ):
                header = {}
                for idx, c in enumerate(low):
                    if ("№" in c or "п/п" in c) and "number" not in header:
                        header["number"] = idx
                    elif "найменув" in c and "name" not in header:
                        header["name"] = idx
                    elif ("один" in c or "вимір" in c) and "unit" not in header:
                        header["unit"] = idx
                    elif ("кільк" in c or "обсяг" in c) and "qty" not in header:
                        header["qty"] = idx
                    elif "примітк" in c and "note" not in header:
                        header["note"] = idx
                continue

        section = _match_section(joined)
        if section is not None:
            current = section
            doc.sections.append(current)
            continue

        if not doc.object_name and _OBJECT_HINT_RE.search(joined) and len(joined) > 25:
            _apply_object_header(doc, joined)
            continue

        if header is None or "name" not in header or "qty" not in header:
            continue

        def cell(key: str) -> str:
            idx = header.get(key, -1)
            return cells[idx] if 0 <= idx < len(cells) else ""

        name, qty = cell("name"), parse_number(cell("qty"))
        if not name or qty is None:
            continue
        if current is None:
            current = ImportedSection(kind="estimate", title="Локальний кошторис")
            doc.sections.append(current)
        counter += 1
        unit_raw = cell("unit")
        current.positions.append(
            ImportedPosition(number=counter, name=clean_text(name),
                             unit=normalize_unit(unit_raw) or clean_text(unit_raw),
                             quantity=qty, note=cell("note"))
        )

    if not doc.positions:
        doc.warnings.append("У файлі не знайдено жодної позиції з обсягом.")
    return doc


def parse_xlsx(data: bytes) -> ImportedDocument:
    return _parse_table_rows(_xlsx_rows(data))


def parse_csv(data: bytes) -> ImportedDocument:
    text = data.decode("utf-8-sig", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=";,\t")
    except csv.Error:
        dialect = csv.excel()
        dialect.delimiter = ";"
    rows = [[clean_text(c) for c in row] for row in csv.reader(io.StringIO(text), dialect)]
    return _parse_table_rows([r for r in rows if any(r)])


# ------------------------------------------------------------------ спільна логіка

def _match_section(line: str) -> ImportedSection | None:
    stripped = clean_text(line)
    m = _ESTIMATE_RE.match(stripped)
    if m:
        rest = clean_text(m.group("rest"))
        code, title = "", ""
        split = re.split(r"\bна\b", rest, maxsplit=1)
        if len(split) == 2:
            code, title = clean_text(split[0]), clean_text(split[1])
        else:
            code = rest
        return ImportedSection(kind="estimate", code=code, title=title)
    m = _DIVISION_RE.match(stripped)
    if m:
        return ImportedSection(kind="division", code=m.group("num") or "",
                               title=clean_text(m.group("title") or ""))
    return None


def _apply_object_header(doc: ImportedDocument, text: str) -> None:
    """Розбирає шапку: назва об'єкта + адреса + місто/область."""
    text = clean_text(text)
    doc.object_name = text
    doc.address = ""
    addr = re.search(r"за\s+адресою:?\s*(.*)$", text, re.I)
    if addr:
        doc.address = clean_text(addr.group(1)).lstrip(": ")
        doc.object_name = clean_text(text[: addr.start()]).rstrip(" ,:")
    source = doc.address or text
    city = re.search(
        r"(?:^|[\s,])(?:м\.\s*|місто\s+|смт\.?\s*|селище\s+міського\s+типу\s+|"
        r"селище\s+|село\s+|с\.\s*)([А-ЯЇІЄҐ][А-Яа-яЇїІіЄєҐґ'\-]+)",
        source,
    )
    if city:
        doc.city = clean_text(city.group(1))
    region = re.search(r"([А-ЯЇІЄҐ][А-Яа-яЇїІіЄєҐґ'\-]+ська)\s+область", source)
    if region:
        doc.region = clean_text(region.group(1)) + " область"


def _finalize_position(block: list[str], number: int, doc: ImportedDocument):
    """Склеює рядки позиції та відрізає хвіст «одиниця виміру + кількість»."""
    text = " ".join(block).rstrip()
    m = _TAIL_STRICT_RE.search(text)
    if m is not None and normalize_unit(m.group("unit")) not in _KNOWN_UNITS_NORM:
        m = None
    if m is None:
        loose = _TAIL_LOOSE_RE.search(text)
        if loose is not None and normalize_unit(loose.group("unit")) in _KNOWN_UNITS_NORM:
            m = loose
    if m is None:
        doc.warnings.append(
            f"Позиція {number}: не розпізнано одиницю виміру/кількість — «{clean_text(text)[:90]}»"
        )
        return None
    qty = parse_number(m.group("qty"))
    if qty is None:
        return None
    return ImportedPosition(number=number, name=clean_text(text[: m.start()]),
                            unit=normalize_unit(m.group("unit")), quantity=qty)


def _parse_lines(lines: list[str]) -> ImportedDocument:
    doc = ImportedDocument()
    current: ImportedSection | None = None
    block: list[str] = []
    block_num = 0
    last_num = 0
    allow_restart = True   # нумерація може починатись з 1 у кожному кошторисі
    header_lines: list[str] = []
    in_header = False
    header_done = False

    def flush():
        nonlocal block, block_num, current
        if block and block_num:
            pos = _finalize_position(block, block_num, doc)
            if pos is not None:
                if current is None:
                    current = ImportedSection(kind="estimate", title="Локальний кошторис")
                    doc.sections.append(current)
                current.positions.append(pos)
        block, block_num = [], 0

    for raw in lines:
        stripped = clean_text(raw)
        if not stripped:
            continue
        if re.match(r"^(Склав|Перевірив|Замовник|Підрядник)\b", stripped, re.I):
            break

        if not doc.doc_code:
            hit = _DOC_CODE_RE.search(stripped)
            if hit:
                doc.doc_code = hit.group(1)

        # --- шапка об'єкта: рядки між «Відомість обсягів робіт» і шапкою таблиці
        if not header_done and _HEADER_TITLE_RE.match(stripped):
            in_header = True
            continue
        if in_header:
            if _TABLE_HEAD_RE.search(stripped) or _match_section(stripped) or \
                    _POSITION_START_RE.match(stripped):
                in_header = False
                header_done = True
                if header_lines:
                    _apply_object_header(doc, " ".join(header_lines))
            else:
                header_lines.append(stripped)
                continue

        section = _match_section(stripped)
        if section is not None:
            flush()
            current = section
            doc.sections.append(current)
            if section.kind == "estimate":
                allow_restart = True
            continue

        if not header_done and not doc.object_name and _OBJECT_HINT_RE.search(stripped) \
                and len(stripped) > 25:
            _apply_object_header(doc, stripped)
            header_done = True
            continue

        m = _POSITION_START_RE.match(raw.lstrip())
        if m:
            num = int(m.group("num"))
            if num == last_num + 1 or (num == 1 and allow_restart):
                flush()
                block_num, last_num = num, num
                allow_restart = False
                block = [m.group("rest")]
                continue

        if block_num:
            block.append(raw.strip() if raw.startswith(" ") is False else raw)
        elif current is not None and current.kind == "estimate" and not current.positions \
                and len(stripped) < 90:
            # Назва локального кошторису перенесена на наступний рядок.
            current.title = clean_text(f"{current.title} {stripped}")

    flush()
    if header_lines and not doc.object_name:
        _apply_object_header(doc, " ".join(header_lines))
    if not doc.positions:
        doc.warnings.append("У файлі не знайдено жодної позиції з обсягом.")
    return doc


# ------------------------------------------------------------------------ точка входу

def parse_document(filename: str, data: bytes) -> ImportedDocument:
    """Визначає формат за розширенням (або сигнатурою) і розбирає відомість."""
    name = (filename or "").lower()
    if name.endswith(".pdf") or data[:4] == b"%PDF":
        return parse_pdf(data)
    if name.endswith((".xlsx", ".xlsm", ".xltx")) or data[:2] == b"PK":
        return parse_xlsx(data)
    return parse_csv(data)
